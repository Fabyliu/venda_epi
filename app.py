from flask import Flask, request, jsonify, render_template_string, send_file
import pandas as pd
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

DEFAULT_CODES = [
    'I06259-M','I06259-G','I06259-GG','I06259-XXG',
    'I06297-M','I06297-G','I06297-GG','I06297-XXG',
    'I06334-M','I06334-G','I06334-GG','I06334-XXG',
    'I06372-M','I06372-G','I06372-GG','I06372-XXG',
    'I06419-M','I06419-G','I06419-GG','I06419-XXG',
    'I06457-M','I06457-G','I06457-GG','I06457-XXG',
    'I06495-M','I06495-G','I06495-GG','I06495-XXG',
    'I06532-M','I06532-G','I06532-GG','I06532-XXG',
    'F301','F302','F303','F305','F307',
    'F401/7','F401/8','F401/9','F401/10',
    'F402/7','F402/8','F402/9','F402/10',
    'F403/7','F403/8','F403/9','F403/10',
    'F404-0','F404-1','F404/2',
    'F405/8','F405/9','F405/10',
    'F406/8','F406/9','F406/10',
    'F407/7','F407/8','F407/9','F407/10',
    'F408/7','F408/8','F408/9','F408/10',
    'F409/7','F409/8','F409/9','F409/10',
    'F410/7','F410/8','F410/9','F410/10',
    'F501/7','F501/8','F501/9','F501/10',
    'F502/7','F502/8','F502/9','F502/10',
    'F503-7','F503-8','F503-9','F503-10',
    'F504-7','F504-8','F504-9','F504-10',
    'F505/8','F505/9','F505/10',
    'F601/7','F601/8','F601/9','F601/10',
    'F602/7','F602/8','F602/9','F602/10',
    'F603/7','F603/8','F603/9','F603/10',
    'F604/7','F604/8','F604/9','F604/10',
    'F605/9-26','F605/9-36','F605/9-46','F605/9-66','F605/10-36',
    'F606/7','F606/8','F606/9','F606/10',
]


def extract_map(file_bytes):
    """Parse Excel file and return {clean_code: quantity} map."""
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    # Find QTD column (QUANTIDADE DO PERÍODO)
    qtd_col = -1
    for i in range(min(5, len(df))):
        for j in range(len(df.columns)):
            h = str(df.iloc[i, j]).replace(' ', '').replace('\xa0', '').upper()
            if 'QUANTIDADEDOPER' in h or ('QUANTIDADE' in h and 'RIO' in h):
                qtd_col = j
                break
        if qtd_col >= 0:
            break
    if qtd_col < 0:
        for i in range(min(5, len(df))):
            for j in range(len(df.columns)):
                h = str(df.iloc[i, j]).replace(' ', '').upper()
                if 'QUANTIDADE' in h and 'TOTAL' in h:
                    qtd_col = j
                    break
            if qtd_col >= 0:
                break
    if qtd_col < 0:
        qtd_col = 7

    m = {}
    for i in range(len(df)):
        val = str(df.iloc[i, 0]).strip()
        if ' - ' in val:
            raw = val.split(' - ')[0].strip()
            clean = raw.replace('INATIVO', '')
            if not clean:
                continue
            try:
                q = float(str(df.iloc[i + 1, qtd_col]).replace('.', '').replace(',', '.'))
            except Exception:
                q = 0.0
            m[clean] = m.get(clean, 0) + q
    return m


def build_results(m_sc, m_sp, codes):
    rows = []
    for c in codes:
        sc = int(round(m_sc.get(c, 0)))
        sp = int(round(m_sp.get(c, 0)))
        rows.append({'code': c, 'sc': sc, 'sp': sp, 'tot': sc + sp})
    return rows


def build_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Consolidado'

    thin = Side(style='thin', color='D0D5DD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', start_color='1F3A5F', end_color='1F3A5F')
    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    alt_fill = PatternFill('solid', start_color='F0F4F8', end_color='F0F4F8')
    tot_fill = PatternFill('solid', start_color='E8F0FE', end_color='E8F0FE')
    data_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', bold=True, size=10)

    headers = ['CÓDIGO', 'QTD SC', 'QTD SP', 'TOTAL SC+SP']
    widths = [18, 16, 16, 18]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    for ri, r in enumerate(rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate([r['code'], r['sc'], r['sp'], r['tot']], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font; cell.border = border
            cell.alignment = left if ci == 1 else right
            if fill: cell.fill = fill
            if ci > 1: cell.number_format = '#,##0'

    tr = len(rows) + 2
    ws.cell(row=tr, column=1, value='TOTAL GERAL').font = bold_font
    ws.cell(row=tr, column=1).fill = tot_fill
    ws.cell(row=tr, column=1).border = border
    ws.cell(row=tr, column=1).alignment = left
    for ci, col in enumerate(['B', 'C', 'D'], 2):
        c = ws.cell(row=tr, column=ci, value=f'=SUM({col}2:{col}{tr-1})')
        c.font = bold_font; c.fill = tot_fill
        c.border = border; c.alignment = right; c.number_format = '#,##0'

    ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


HTML = r"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Vendas por Código — SC + SP</title>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=Bebas+Neue&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#f5f2ed;--paper:#faf8f4;--ink:#1a1a18;--ink2:#4a4a45;--ink3:#9a9a90;
  --rule:#ddd9d0;--accent:#c8402a;--green:#2a7a4a;--blue:#1a4a8a;--yellow:#f5d04a;
  --mono:'IBM Plex Mono',monospace;--display:'Bebas Neue',sans-serif;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{font-size:15px}
body{background:var(--bg);color:var(--ink);font-family:var(--mono);min-height:100vh}
body::before{content:'';position:fixed;inset:0;
  background-image:radial-gradient(circle,rgba(26,26,24,.08) 1px,transparent 1px);
  background-size:24px 24px;pointer-events:none;z-index:0}
.wrap{position:relative;z-index:1;max-width:960px;margin:0 auto;padding:2.5rem 2rem 5rem}

/* HEADER */
.header{border-top:3px solid var(--ink);border-bottom:1px solid var(--rule);
  padding:1.5rem 0 1.25rem;margin-bottom:2rem;
  display:flex;align-items:flex-end;justify-content:space-between;flex-wrap:wrap;gap:12px}
.issue-line{font-size:10px;letter-spacing:.18em;color:var(--ink3);text-transform:uppercase;margin-bottom:6px}
h1{font-family:var(--display);font-size:clamp(2.4rem,6vw,4rem);letter-spacing:.03em;line-height:1}
h1 em{color:var(--accent);font-style:normal}
.edition-box{border:1px solid var(--ink);padding:6px 12px;font-size:10px;
  letter-spacing:.1em;text-transform:uppercase;color:var(--ink2);line-height:1.8}

/* SECTION LABEL */
.slabel{font-size:9px;letter-spacing:.2em;text-transform:uppercase;color:var(--ink3);
  margin-bottom:8px;display:flex;align-items:center;gap:8px}
.slabel::after{content:'';flex:1;height:1px;background:var(--rule)}

/* UPLOAD */
.upload-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px}
@media(max-width:520px){.upload-grid{grid-template-columns:1fr}}
.slot{background:var(--paper);border:1px solid var(--rule);padding:14px 16px;
  cursor:pointer;position:relative;transition:border-color .15s}
.slot:hover{border-color:var(--ink2)}
.slot.loaded{border-color:var(--ink);border-left:3px solid var(--green)}
.slot input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.slot-tag{font-size:9px;letter-spacing:.15em;text-transform:uppercase;color:var(--accent);margin-bottom:5px}
.slot-name{font-size:12px;font-weight:600;color:var(--ink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.slot-meta{font-size:10px;color:var(--ink3);margin-top:2px}
.slot-check{position:absolute;top:10px;right:12px;font-size:10px;color:var(--green);
  opacity:0;transition:opacity .2s;font-weight:600}
.slot.loaded .slot-check{opacity:1}
.drop-bar{border:1px dashed var(--rule);background:var(--paper);padding:12px;
  text-align:center;position:relative;cursor:pointer;transition:border-color .15s}
.drop-bar:hover,.drop-bar.over{border-color:var(--ink2)}
.drop-bar input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.drop-bar span{font-size:11px;color:var(--ink3)}

/* BUTTONS */
.run-btn{width:100%;margin-top:12px;padding:13px;background:var(--ink);color:var(--bg);
  font-family:var(--display);font-size:1.1rem;letter-spacing:.12em;border:none;
  cursor:pointer;transition:background .15s,transform .1s;
  display:flex;align-items:center;justify-content:center;gap:10px}
.run-btn:hover:not(:disabled){background:var(--accent)}
.run-btn:active:not(:disabled){transform:scale(.99)}
.run-btn:disabled{opacity:.3;cursor:default}
.prog-wrap{height:2px;background:var(--rule);margin-top:8px;display:none}
.prog-bar{height:100%;background:var(--accent);width:0;transition:width .3s}
.status-line{font-size:11px;color:var(--ink3);min-height:18px;margin-top:6px;text-align:center}
.status-line.ok{color:var(--green);font-weight:600}
.status-line.err{color:var(--accent)}

/* CODE MANAGER */
.mgr-toggle{display:flex;align-items:center;gap:10px;margin-top:1.5rem}
.mgr-toggle-btn{padding:5px 12px;border:1px solid var(--rule);background:var(--paper);
  color:var(--ink2);font-family:var(--mono);font-size:10px;letter-spacing:.08em;
  cursor:pointer;text-transform:uppercase;transition:all .15s;white-space:nowrap}
.mgr-toggle-btn:hover{border-color:var(--ink);color:var(--ink)}
.mgr-toggle-btn.open{background:var(--ink);color:var(--bg);border-color:var(--ink)}
.mgr-panel{display:none;margin-top:10px;border:1px solid var(--rule);background:var(--paper)}
.mgr-panel.open{display:block}
.mgr-inner{padding:16px}
.add-row{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap;align-items:center}
.add-input{flex:1;min-width:140px;padding:8px 10px;border:1px solid var(--rule);
  background:var(--bg);color:var(--ink);font-family:var(--mono);font-size:12px;
  outline:none;transition:border-color .15s;text-transform:uppercase}
.add-input:focus{border-color:var(--ink)}
.add-input::placeholder{color:var(--ink3);text-transform:none}
.add-btn{padding:8px 14px;border:1px solid var(--ink);background:var(--ink);color:var(--bg);
  font-family:var(--mono);font-size:11px;cursor:pointer;text-transform:uppercase;
  transition:all .15s;white-space:nowrap}
.add-btn:hover{background:var(--accent);border-color:var(--accent)}
.add-msg{font-size:11px;width:100%;margin-top:-6px}
.add-msg.ok{color:var(--green)}.add-msg.err{color:var(--accent)}
.code-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(130px,1fr));
  gap:6px;max-height:280px;overflow-y:auto;padding:2px}
.code-grid::-webkit-scrollbar{width:4px}
.code-grid::-webkit-scrollbar-thumb{background:var(--rule)}
.code-tag{display:flex;align-items:center;justify-content:space-between;
  padding:5px 8px;border:1px solid var(--rule);background:var(--bg);font-size:11px;font-weight:500}
.code-tag.new-tag{border-color:var(--green);color:var(--green)}
.del-btn{margin-left:6px;cursor:pointer;color:var(--ink3);font-size:14px;line-height:1;
  background:none;border:none;font-family:var(--mono);padding:0 2px;transition:color .12s}
.del-btn:hover{color:var(--accent)}
.mgr-footer{display:flex;align-items:center;justify-content:space-between;
  margin-top:12px;padding-top:12px;border-top:1px solid var(--rule);flex-wrap:wrap;gap:8px}
.mgr-count{font-size:11px;color:var(--ink3)}
.mgr-count strong{color:var(--ink)}
.mgr-actions{display:flex;gap:6px}
.mgr-act{padding:5px 10px;border:1px solid var(--rule);background:transparent;
  color:var(--ink2);font-family:var(--mono);font-size:10px;cursor:pointer;
  text-transform:uppercase;transition:all .12s;letter-spacing:.04em}
.mgr-act:hover{border-color:var(--ink);color:var(--ink)}

/* RESULTS */
.results{margin-top:2.5rem;display:none}
.stats-band{border-top:2px solid var(--ink);border-bottom:1px solid var(--rule);
  display:grid;grid-template-columns:repeat(4,1fr);margin-bottom:1.5rem}
@media(max-width:540px){.stats-band{grid-template-columns:1fr 1fr}}
.stat{padding:14px 16px;border-right:1px solid var(--rule)}
.stat:last-child{border-right:none}
.stat-num{font-family:var(--display);font-size:2rem;line-height:1;color:var(--ink)}
.stat-num.accent{color:var(--accent)}
.stat-lbl{font-size:9px;letter-spacing:.12em;text-transform:uppercase;color:var(--ink3);margin-top:3px}
.toolbar{display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap;align-items:center}
.search-box{flex:1;min-width:140px;padding:7px 10px;border:1px solid var(--rule);
  background:var(--paper);color:var(--ink);font-family:var(--mono);font-size:11px;
  outline:none;transition:border-color .15s}
.search-box:focus{border-color:var(--ink)}
.search-box::placeholder{color:var(--ink3)}
.chip{padding:7px 12px;border:1px solid var(--rule);background:var(--paper);
  color:var(--ink3);font-family:var(--mono);font-size:10px;letter-spacing:.08em;
  cursor:pointer;transition:all .12s;text-transform:uppercase}
.chip:hover{border-color:var(--ink2);color:var(--ink)}
.chip.active{background:var(--ink);color:var(--bg);border-color:var(--ink)}
.dl-btn{padding:7px 12px;border:1px solid var(--rule);background:var(--paper);
  color:var(--ink2);font-family:var(--mono);font-size:10px;cursor:pointer;
  transition:all .12s;display:flex;align-items:center;gap:5px;
  text-transform:uppercase;letter-spacing:.06em;text-decoration:none}
.dl-btn:hover{border-color:var(--ink);color:var(--ink)}
.tbl-outer{border:1px solid var(--rule);background:var(--paper)}
.tbl-scroll{max-height:460px;overflow-y:auto}
.tbl-scroll::-webkit-scrollbar{width:5px}
.tbl-scroll::-webkit-scrollbar-thumb{background:var(--rule)}
table{width:100%;border-collapse:collapse;font-size:12px}
thead th{position:sticky;top:0;padding:9px 14px;background:var(--ink);color:var(--bg);
  font-size:9px;letter-spacing:.15em;text-transform:uppercase;font-weight:500;text-align:left}
thead th:not(:first-child){text-align:right}
tbody tr{border-bottom:1px solid var(--rule);transition:background .08s}
tbody tr:hover{background:rgba(200,64,42,.04)}
tbody tr.odd{background:rgba(26,26,24,.02)}
tbody tr.odd:hover{background:rgba(200,64,42,.05)}
tbody tr.zero{opacity:.35}
tbody td{padding:7px 14px}
tbody td:first-child{font-weight:500}
tbody td:not(:first-child){text-align:right;font-variant-numeric:tabular-nums}
.sc{color:var(--blue)}.sp{color:var(--green)}.tot{font-weight:600}
.zero-val{color:var(--ink3)}
.grp-hdr td{background:var(--bg);font-size:9px;letter-spacing:.15em;text-transform:uppercase;
  color:var(--ink3);padding:5px 14px;border-bottom:1px solid var(--rule)}
tfoot td{padding:10px 14px;border-top:2px solid var(--ink);background:var(--ink);
  color:var(--bg);font-weight:600;font-size:12px}
tfoot td:not(:first-child){text-align:right}
.grand{font-family:var(--display);font-size:1.15rem;color:var(--yellow)}
.footer{margin-top:2.5rem;border-top:1px solid var(--rule);padding-top:1rem;
  display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.footer-brand{font-family:var(--display);font-size:1rem;letter-spacing:.1em;color:var(--ink3)}
.footer-note{font-size:10px;color:var(--ink3)}
@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
.results{animation:fadeIn .3s ease forwards}
</style>
</head>
<body>
<div class="wrap">
  <header class="header">
    <div>
      <div class="issue-line">Relatório de Vendas · SC + SP</div>
      <h1>VENDAS<br><em>POR CÓDIGO</em></h1>
    </div>
    <div class="edition-box" id="edBox">— CÓDIGOS<br>QUANTIDADE DO PERÍODO<br>CONSOLIDAÇÃO SC + SP</div>
  </header>

  <div class="slabel">Carregar arquivos</div>
  <div class="upload-grid">
    <div class="slot" id="s0"><input type="file" accept=".xlsx" id="f0">
      <div class="slot-tag">Arquivo SC</div>
      <div class="slot-name" id="n0">vendas-SC...xlsx</div>
      <div class="slot-meta" id="m0">Clique para selecionar</div>
      <span class="slot-check">✓ CARREGADO</span>
    </div>
    <div class="slot" id="s1"><input type="file" accept=".xlsx" id="f1">
      <div class="slot-tag">Arquivo SP</div>
      <div class="slot-name" id="n1">vendas-SP...xlsx</div>
      <div class="slot-meta" id="m1">Clique para selecionar</div>
      <span class="slot-check">✓ CARREGADO</span>
    </div>
  </div>
  <div class="drop-bar" id="drop">
    <input type="file" accept=".xlsx" multiple id="fmulti">
    <span>或將兩個 .xlsx 拖曳至此</span>
  </div>
  <div class="prog-wrap" id="prog"><div class="prog-bar" id="pbar"></div></div>
  <div class="status-line" id="status"></div>
  <button class="run-btn" id="runBtn" disabled onclick="analyse()">
    <svg width="15" height="15" viewBox="0 0 24 24" fill="currentColor"><polygon points="5 3 19 12 5 21"/></svg>
    ANALISAR E CONSOLIDAR
  </button>

  <!-- CODE MANAGER -->
  <div class="mgr-toggle">
    <div class="slabel" style="flex:1;margin:0;cursor:pointer" onclick="toggleMgr()">Gerenciar códigos</div>
    <button class="mgr-toggle-btn" id="mgrBtn" onclick="toggleMgr()">▸ EDITAR LISTA</button>
  </div>
  <div class="mgr-panel" id="mgrPanel">
    <div class="mgr-inner">
      <div class="add-row">
        <input class="add-input" id="addIn" placeholder="Novo código, ex: F411/9"
          oninput="this.value=this.value.toUpperCase()" onkeydown="if(event.key==='Enter')addCode()">
        <button class="add-btn" onclick="addCode()">+ Adicionar</button>
        <div class="add-msg" id="addMsg"></div>
      </div>
      <div class="code-grid" id="codeGrid"></div>
      <div class="mgr-footer">
        <span class="mgr-count">Total: <strong id="codeCount">0</strong> códigos</span>
        <div class="mgr-actions">
          <button class="mgr-act" onclick="resetCodes()">↺ Padrão</button>
          <button class="mgr-act" onclick="exportCodes()">↓ Exportar</button>
          <button class="mgr-act" onclick="document.getElementById('impIn').click()">↑ Importar</button>
          <input type="file" id="impIn" accept=".txt,.csv" style="display:none" onchange="importCodes(this)">
        </div>
      </div>
    </div>
  </div>

  <!-- RESULTS -->
  <div class="results" id="results">
    <div class="stats-band" id="statsRow"></div>
    <div class="toolbar">
      <input class="search-box" id="sq" placeholder="Buscar código..." oninput="filter()">
      <button class="chip active" id="cAll" onclick="setF('all')">Todos</button>
      <button class="chip" id="cF" onclick="setF('F')">Série F</button>
      <button class="chip" id="cI" onclick="setF('I')">Série I</button>
      <button class="chip" id="cZ" onclick="setF('zero')">Sem Qty</button>
      <a class="dl-btn" id="dlCSV" href="#">
        <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
        CSV
      </a>
      <a class="dl-btn" id="dlXLSX" href="#">
        <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>
        XLSX
      </a>
    </div>
    <div class="tbl-outer"><div class="tbl-scroll">
      <table>
        <thead><tr><th>Código</th><th>Qtd SC</th><th>Qtd SP</th><th>Total SC+SP</th></tr></thead>
        <tbody id="tbody"></tbody>
        <tfoot><tr>
          <td>TOTAL GERAL</td>
          <td id="ftSC">—</td><td id="ftSP">—</td><td id="ftTot" class="grand">—</td>
        </tr></tfoot>
      </table>
    </div></div>
  </div>

  <footer class="footer">
    <span class="footer-brand">VC · TOOL</span>
    <span class="footer-note">Análise feita no servidor — precisão garantida pelo Python.</span>
  </footer>
</div>

<script>
const DEFAULT_CODES = """ + json.dumps(DEFAULT_CODES) + r""";
const fmt = n => n === 0 ? '—' : n.toLocaleString('pt-BR');

let userCodes = JSON.parse(localStorage.getItem('vc_codes') || 'null') || [...DEFAULT_CODES];
let newCodes = new Set();
let allRows = [];
let activeF = 'all';
let lastJobId = null;

function saveCodes(){ localStorage.setItem('vc_codes', JSON.stringify(userCodes)); }

// ── FILES ──
let files = [null, null];
function pickFile(i, input) {
  const f = input.files[0]; if (!f) return;
  files[i] = f;
  document.getElementById('n'+i).textContent = f.name;
  document.getElementById('m'+i).textContent = (f.size/1024).toFixed(1)+' KB';
  document.getElementById('s'+i).classList.add('loaded');
  checkReady();
}
document.getElementById('f0').onchange = e => pickFile(0, e.target);
document.getElementById('f1').onchange = e => pickFile(1, e.target);
document.getElementById('fmulti').onchange = function() {
  Array.from(this.files).filter(f=>f.name.endsWith('.xlsx')).slice(0,2).forEach((f,i)=>{
    files[i]=f;
    document.getElementById('n'+i).textContent=f.name;
    document.getElementById('m'+i).textContent=(f.size/1024).toFixed(1)+' KB';
    document.getElementById('s'+i).classList.add('loaded');
  }); checkReady();
};
const drop = document.getElementById('drop');
drop.ondragover = e => { e.preventDefault(); drop.classList.add('over'); };
drop.ondragleave = () => drop.classList.remove('over');
drop.ondrop = e => {
  e.preventDefault(); drop.classList.remove('over');
  Array.from(e.dataTransfer.files).filter(f=>f.name.endsWith('.xlsx')).slice(0,2).forEach((f,i)=>{
    files[i]=f;
    document.getElementById('n'+i).textContent=f.name;
    document.getElementById('m'+i).textContent=(f.size/1024).toFixed(1)+' KB';
    document.getElementById('s'+i).classList.add('loaded');
  }); checkReady();
};
function checkReady(){
  document.getElementById('runBtn').disabled = !(files[0]&&files[1]);
  if(files[0]&&files[1]) setStatus('兩個檔案已就緒','ok');
}
function setStatus(m,c=''){
  const el=document.getElementById('status'); el.textContent=m; el.className='status-line'+(c?' '+c:'');
}

// ── ANALYSE ──
async function analyse() {
  setStatus('上傳並分析中...');
  const prog = document.getElementById('prog'), bar = document.getElementById('pbar');
  prog.style.display='block'; bar.style.width='20%';
  document.getElementById('runBtn').disabled = true;

  try {
    const fd = new FormData();
    fd.append('sc', files[0]);
    fd.append('sp', files[1]);
    fd.append('codes', JSON.stringify(userCodes));

    bar.style.width = '60%';
    const res = await fetch('/analyse', { method:'POST', body:fd });
    const data = await res.json();

    bar.style.width = '100%';
    setTimeout(() => prog.style.display='none', 300);

    if (data.error) {
      setStatus('錯誤：' + data.error, 'err');
    } else {
      allRows = data.rows;
      lastJobId = data.job_id;
      renderStats();
      filter();
      document.getElementById('results').style.display = 'block';
      document.getElementById('dlCSV').href = '/download/csv/' + lastJobId;
      document.getElementById('dlXLSX').href = '/download/xlsx/' + lastJobId;
      setStatus('✓ 完成 — ' + allRows.length + ' 個代碼', 'ok');
      document.getElementById('edBox').innerHTML = allRows.length + ' CÓDIGOS<br>QUANTIDADE DO PERÍODO<br>CONSOLIDAÇÃO SC + SP';
    }
  } catch(e) {
    setStatus('網路錯誤：' + e.message, 'err');
    prog.style.display = 'none';
  }
  document.getElementById('runBtn').disabled = false;
}

// ── RENDER ──
function renderStats() {
  const tSC=allRows.reduce((s,r)=>s+r.sc,0);
  const tSP=allRows.reduce((s,r)=>s+r.sp,0);
  document.getElementById('statsRow').innerHTML = [
    {n:allRows.length,l:'Códigos'},
    {n:tSC.toLocaleString('pt-BR'),l:'Total SC'},
    {n:tSP.toLocaleString('pt-BR'),l:'Total SP'},
    {n:(tSC+tSP).toLocaleString('pt-BR'),l:'Grand Total',a:true},
  ].map(s=>`<div class="stat"><div class="stat-num${s.a?' accent':''}">${s.n}</div><div class="stat-lbl">${s.l}</div></div>`).join('');
}
function getGroup(c){ return c.startsWith('I') ? c.slice(0,6) : (c.match(/^(F\d+)/)||[,'?'])[1]; }
function setF(f){
  activeF=f;
  ['All','F','I','Z'].forEach(k=>document.getElementById('c'+k).classList.toggle('active',f===(k==='All'?'all':k==='Z'?'zero':k)));
  filter();
}
function filter(){
  const q=document.getElementById('sq').value.toUpperCase();
  renderTable(allRows.filter(r=>{
    if(q&&!r.code.includes(q)) return false;
    if(activeF==='F') return r.code.startsWith('F');
    if(activeF==='I') return r.code.startsWith('I');
    if(activeF==='zero') return r.tot===0;
    return true;
  }));
}
function renderTable(data){
  let html='',prev='',tog=false;
  data.forEach(r=>{
    const g=getGroup(r.code);
    if(g!==prev){prev=g;tog=!tog;html+=`<tr class="grp-hdr"><td colspan="4">Série ${g}</td></tr>`;}
    const rc=(tog?'odd':'')+(r.tot===0?' zero':'');
    const sc=r.sc===0?`<span class="zero-val">—</span>`:`<span class="sc">${r.sc.toLocaleString('pt-BR')}</span>`;
    const sp=r.sp===0?`<span class="zero-val">—</span>`:`<span class="sp">${r.sp.toLocaleString('pt-BR')}</span>`;
    const tot=r.tot===0?`<span class="zero-val">—</span>`:`<span class="tot">${r.tot.toLocaleString('pt-BR')}</span>`;
    html+=`<tr class="${rc}"><td>${r.code}</td><td>${sc}</td><td>${sp}</td><td>${tot}</td></tr>`;
  });
  document.getElementById('tbody').innerHTML=html;
  const tSC=data.reduce((s,r)=>s+r.sc,0),tSP=data.reduce((s,r)=>s+r.sp,0);
  document.getElementById('ftSC').textContent=tSC.toLocaleString('pt-BR');
  document.getElementById('ftSP').textContent=tSP.toLocaleString('pt-BR');
  document.getElementById('ftTot').textContent=(tSC+tSP).toLocaleString('pt-BR');
}

// ── CODE MANAGER ──
function toggleMgr(){
  const p=document.getElementById('mgrPanel'),b=document.getElementById('mgrBtn');
  const open=p.classList.toggle('open');
  b.classList.toggle('open',open);
  b.textContent=open?'▾ FECHAR':'▸ EDITAR LISTA';
  if(open) renderGrid();
}
function renderGrid(){
  document.getElementById('codeGrid').innerHTML = userCodes.map((c,i)=>`
    <div class="code-tag${newCodes.has(c)?' new-tag':''}">
      <span>${c}</span>
      <button class="del-btn" onclick="delCode(${i})">×</button>
    </div>`).join('');
  document.getElementById('codeCount').textContent = userCodes.length;
}
function addCode(){
  const inp=document.getElementById('addIn'),msg=document.getElementById('addMsg');
  const v=inp.value.trim().toUpperCase();
  if(!v){showMsg(msg,'Por favor insira um código','err');return;}
  if(userCodes.includes(v)){showMsg(msg,`"${v}" já existe`,'err');return;}
  userCodes.push(v); newCodes.add(v); saveCodes(); renderGrid(); inp.value='';
  showMsg(msg,`"${v}" adicionado`,'ok');
}
function delCode(i){
  if(!confirm(`Remover "${userCodes[i]}"?`)) return;
  newCodes.delete(userCodes[i]); userCodes.splice(i,1); saveCodes(); renderGrid();
}
function resetCodes(){
  if(!confirm('Repor lista padrão?')) return;
  userCodes=[...DEFAULT_CODES]; newCodes.clear(); saveCodes(); renderGrid();
  showMsg(document.getElementById('addMsg'),'Lista reposta','ok');
}
function exportCodes(){
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([userCodes.join('\n')],{type:'text/plain'}));
  a.download='codigos.txt'; a.click();
}
function importCodes(input){
  const r=new FileReader();
  r.onload=e=>{
    const lines=e.target.result.split(/[\r\n,;]+/).map(l=>l.trim().toUpperCase()).filter(Boolean);
    let n=0; lines.forEach(c=>{if(!userCodes.includes(c)){userCodes.push(c);newCodes.add(c);n++;}});
    saveCodes(); renderGrid(); showMsg(document.getElementById('addMsg'),`${n} importado(s)`,'ok');
  };
  r.readAsText(input.files[0]); input.value='';
}
function showMsg(el,txt,type){
  el.textContent=txt; el.className='add-msg '+type;
  setTimeout(()=>{el.textContent='';el.className='add-msg';},3000);
}

// ── INIT ──
document.getElementById('edBox').innerHTML = userCodes.length + ' CÓDIGOS<br>QUANTIDADE DO PERÍODO<br>CONSOLIDAÇÃO SC + SP';
</script>
</body>
</html>
"""


# In-memory job store (for download links)
import uuid
job_store = {}


@app.route('/')
def index():
    return render_template_string(HTML)


@app.route('/analyse', methods=['POST'])
def analyse():
    try:
        sc_bytes = request.files['sc'].read()
        sp_bytes = request.files['sp'].read()
        codes = json.loads(request.form.get('codes', 'null')) or DEFAULT_CODES

        m_sc = extract_map(sc_bytes)
        m_sp = extract_map(sp_bytes)
        rows = build_results(m_sc, m_sp, codes)

        job_id = str(uuid.uuid4())
        job_store[job_id] = {'rows': rows, 'sc': sc_bytes, 'sp': sp_bytes, 'codes': codes}

        return jsonify({'rows': rows, 'job_id': job_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/download/csv/<job_id>')
def download_csv(job_id):
    job = job_store.get(job_id)
    if not job:
        return 'Not found', 404
    rows = job['rows']
    lines = ['\ufeffCÓDIGO,QTD SC,QTD SP,TOTAL SC+SP']
    for r in rows:
        lines.append(f"{r['code']},{r['sc']},{r['sp']},{r['tot']}")
    tsc = sum(r['sc'] for r in rows)
    tsp = sum(r['sp'] for r in rows)
    lines.append(f"TOTAL GERAL,{tsc},{tsp},{tsc+tsp}")
    buf = io.BytesIO('\n'.join(lines).encode('utf-8-sig'))
    buf.seek(0)
    return send_file(buf, mimetype='text/csv',
                     as_attachment=True, download_name='vendas_por_codigo.csv')


@app.route('/download/xlsx/<job_id>')
def download_xlsx(job_id):
    job = job_store.get(job_id)
    if not job:
        return 'Not found', 404
    buf = build_xlsx(job['rows'])
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='vendas_por_codigo.xlsx')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
